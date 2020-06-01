#https://stackoverflow.com/questions/56880675/how-to-read-merged-cells-in-python-using-openpyxl
#openpyxl doesnot read merge cells
from os import chdir, listdir
from copy import copy
import xlrd
from openpyxl import load_workbook, Workbook

def createNewWorkbook(manyWb):
    for wb in manyWb:
        for sheetName in wb.sheet_names():
            o = theOne.create_sheet(sheetName)
            safeTitle = o.title
            copySheet(wb[sheetName],theOne[safeTitle])

def copySheet(sourceSheet,newSheet):
    for row in sourceSheet.rows:
        for cell in row:
            newCell = newSheet.cell(row=cell.row, column= cell.col_idx, value= cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)


if __name__=="__main__":
    inputLocation = "C:/MergeExcel/Input"
    outputLocation = "C:/MergeExcel/Output"
    chdir(inputLocation)
    theOneFile = "Combine.xlsx"
    myxlrd = [xlrd.open_workbook(f) for f in listdir('.')]
    print(myxlrd)
    # try this if you are bored
    # myfriends = [ openpyxl.load_workbook(f) for k in range(200) for f in filesInput ]

    theOne = Workbook()
    del theOne['Sheet']  # We want our new book to be empty. Thanks.
    createNewWorkbook(myxlrd)
    chdir(outputLocation)
    theOne.save(theOneFile)